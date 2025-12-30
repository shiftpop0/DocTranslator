# translate/excel.py
import datetime
import logging
from typing import List, Dict, Any
from threading import Event
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from . import to_translate
from . import common


def start(trans: Dict[str, Any]) -> bool:
    """
    Excel文件翻译入口
    :param trans: 翻译配置字典
    :return: 是否成功
    """
    translate_id = trans['id']
    start_time = datetime.datetime.now()

    # 加载工作簿
    try:
        wb = openpyxl.load_workbook(trans['file_path'])
    except Exception as e:
        logging.error(f"[任务{translate_id}] 无法打开Excel文件: {e}")
        to_translate.error(translate_id, f"无法打开Excel文件: {str(e)}")
        return False

    # 提取所有需要翻译的单元格
    texts = []
    cell_map = []  # 记录单元格位置，用于回写

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _extract_sheet_texts(ws, sheet_name, texts, cell_map)
    except Exception as e:
        logging.error(f"[任务{translate_id}] 提取文本失败: {e}")
        to_translate.error(translate_id, f"提取文本失败: {str(e)}")
        return False

    if not texts:
        logging.info(f"[任务{translate_id}] Excel中没有需要翻译的内容")
        wb.save(trans['target_file'])
        to_translate.complete(trans, 0, "0秒")
        return True

    logging.info(f"[任务{translate_id}] 提取到 {len(texts)} 个单元格需要翻译")

    # 【关键修改】使用线程池批量翻译
    event = Event()
    success = to_translate.translate_batch(trans, texts, event)
    if not success:
        return False

    # 回写翻译结果
    try:
        text_count = _apply_translation(wb, texts, cell_map, trans.get('type', ''))
        wb.save(trans['target_file'])
    except Exception as e:
        logging.error(f"[任务{translate_id}] 保存文件失败: {e}")
        to_translate.error(translate_id, f"保存文件失败: {str(e)}")
        return False

    end_time = datetime.datetime.now()
    spend_time = common.display_spend(start_time, end_time)
    to_translate.complete(trans, text_count, spend_time)
    return True


def _extract_sheet_texts(ws: Worksheet, sheet_name: str, texts: List[Dict], cell_map: List[Dict]):
    """
    提取工作表中的文本
    :param ws: 工作表对象
    :param sheet_name: 工作表名称
    :param texts: 文本列表（输出）
    :param cell_map: 单元格位置映射（输出）
    """
    # 获取合并单元格范围，避免重复翻译
    merged_cells = set()
    for merged_range in ws.merged_cells.ranges:
        # 只保留合并区域的第一个单元格，其他跳过
        first_cell = True
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if first_cell:
                    first_cell = False
                else:
                    merged_cells.add((row, col))

    # 遍历所有单元格
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            # 跳过合并单元格的非首单元格
            if (row_idx, col_idx) in merged_cells:
                continue

            value = cell.value
            if _should_translate(value):
                text_item = {
                    'text': str(value),
                    'original': str(value),
                    'complete': False,
                    'count': 0
                }
                texts.append(text_item)
                cell_map.append({
                    'sheet': sheet_name,
                    'row': row_idx,
                    'col': col_idx,
                    'text_index': len(texts) - 1
                })


def _should_translate(value) -> bool:
    """判断单元格值是否需要翻译"""
    if value is None:
        return False
    if isinstance(value, (int, float, complex)):
        return False
    if isinstance(value, datetime.datetime):
        return False
    if isinstance(value, datetime.time):
        return False

    text = str(value).strip()
    if not text:
        return False
    if common.is_all_punc(text):
        return False

    return True



def _apply_translation(wb, texts: List[Dict], cell_map: List[Dict], trans_type: str) -> int:
    """
    应用翻译结果到工作簿
    :return: 翻译字数统计
    """
    text_count = 0
    keep_both = 'both' in trans_type

    for mapping in cell_map:
        sheet_name = mapping['sheet']
        row = mapping['row']
        col = mapping['col']
        text_index = mapping['text_index']

        if text_index >= len(texts):
            continue

        text_item = texts[text_index]
        text_count += text_item.get('count', 0)

        ws = wb[sheet_name]
        cell = ws.cell(row=row, column=col)

        original = text_item.get('original', '')
        translated = text_item.get('text', original)

        if keep_both:
            # 保留原文和译文，用换行分隔
            cell.value = f"{original}\n{translated}"
        else:
            # 仅保留译文
            cell.value = translated

    return text_count
